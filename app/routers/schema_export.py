"""
테이블 정의서 엑셀 내보내기 API
"""
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse, JSONResponse
from sqlalchemy.orm import Session
from typing import Optional, List
from datetime import datetime
from io import BytesIO
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.core.database import get_db, User
from app.services.server_service import ServerService
from app.services.drivers import get_driver
from app.routers.auth import require_login

from urllib.parse import quote
from app.services.activity_service import log_download_schema, log_download_schema_all

router = APIRouter(prefix="/api/schema-export", tags=["schema-export"])

# 스타일 정의
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
SUBHEADER_FONT = Font(bold=True, size=10)
TITLE_FONT = Font(bold=True, size=14)
SUBTITLE_FONT = Font(bold=True, size=12)
NORMAL_FONT = Font(size=10)
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
LEFT_ALIGN = Alignment(horizontal='left', vertical='center')


def set_column_widths(ws, widths: dict):
    """컬럼 너비 설정"""
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def create_db_toc_sheet(wb, server_name: str, db_name: str, tables: list, user_name: str):
    """단일 DB용 목차 시트 생성"""
    ws = wb.active
    ws.title = "목차"
    
    # 제목
    ws.merge_cells('A1:H1')
    ws['A1'] = "테이블 정의서"
    ws['A1'].font = Font(bold=True, size=20)
    ws['A1'].alignment = CENTER_ALIGN
    ws.row_dimensions[1].height = 40
    
    # 정보
    ws['A3'] = "서버"
    ws['B3'] = server_name
    ws['A3'].font = Font(bold=True)
    
    ws['A4'] = "DB명"
    ws['B4'] = db_name
    ws['A4'].font = Font(bold=True)
    
    ws['A5'] = "생성일"
    ws['B5'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws['A5'].font = Font(bold=True)
    
    ws['A6'] = "생성자"
    ws['B6'] = user_name
    ws['A6'].font = Font(bold=True)
    
    # 요약
    ws['D3'] = "테이블 수"
    ws['E3'] = len(tables)
    ws['D3'].font = Font(bold=True)
    
    total_rows = sum(t.get('row_count', 0) for t in tables)
    ws['D4'] = "총 행 수"
    ws['E4'] = total_rows
    ws['D4'].font = Font(bold=True)
    
    total_size = sum(t.get('size_mb', 0) for t in tables)
    ws['D5'] = "용량(MB)"
    ws['E5'] = round(total_size, 2)
    ws['D5'].font = Font(bold=True)
    
    # 테이블 목록
    ws['A8'] = "■ 테이블 목록"
    ws['A8'].font = SUBTITLE_FONT
    
    headers = ['No', '테이블명', '테이블설명', '컬럼 수', '행 수', '용량(MB)', '시트명']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=9, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = BORDER
    
    for row_idx, table in enumerate(tables, 10):
        sheet_name = table['table_name'][:31]
        table_desc = table.get('description') or table.get('table_description') or ''
        
        ws.cell(row=row_idx, column=1, value=row_idx - 9).border = BORDER
        ws.cell(row=row_idx, column=1).alignment = CENTER_ALIGN
        ws.cell(row=row_idx, column=2, value=table['table_name']).border = BORDER
        ws.cell(row=row_idx, column=3, value=table_desc).border = BORDER
        ws.cell(row=row_idx, column=4, value=table.get('column_count', 0)).border = BORDER
        ws.cell(row=row_idx, column=4).alignment = CENTER_ALIGN
        ws.cell(row=row_idx, column=5, value=table.get('row_count', 0)).border = BORDER
        ws.cell(row=row_idx, column=5).alignment = CENTER_ALIGN
        ws.cell(row=row_idx, column=6, value=round(table.get('size_mb', 0), 2)).border = BORDER
        ws.cell(row=row_idx, column=6).alignment = CENTER_ALIGN
        
        # 시트 링크
        cell = ws.cell(row=row_idx, column=7, value=sheet_name)
        cell.hyperlink = f"#'{sheet_name}'!A1"
        cell.font = Font(color="0563C1", underline="single")
        cell.border = BORDER
    
    set_column_widths(ws, {'A': 8, 'B': 25, 'C': 30, 'D': 10, 'E': 12, 'F': 12, 'G': 25})


def create_table_sheet(wb, db_name: str, table_name: str, columns: list, table_info: dict):
    """테이블 정의 시트 생성"""
    # 시트명 (31자 제한, 특수문자 제거)
    sheet_name = table_name[:31].replace('/', '_').replace('\\', '_').replace('*', '_')
    ws = wb.create_sheet(title=sheet_name)
    
    # 테이블 정보
    table_desc = table_info.get('description') or table_info.get('table_description') or ''
    if table_desc:
        ws['A1'] = f"테이블명: {table_name} ({table_desc})"
    else:
        ws['A1'] = f"테이블명: {table_name}"
    ws['A1'].font = SUBTITLE_FONT
    
    ws['A2'] = f"행 수: {table_info.get('row_count', 0):,} | 용량: {round(table_info.get('size_mb', 0), 2)} MB"
    ws['A2'].font = Font(size=10, color="666666")
    
    # 컬럼 헤더
    headers = ['No', '컬럼명', '데이터타입', '길이', 'PK', 'NULL', '기본값', '설명', '비고']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = BORDER
    
    # 컬럼 데이터
    for row_idx, col in enumerate(columns, 5):
        ws.cell(row=row_idx, column=1, value=row_idx - 4).border = BORDER
        ws.cell(row=row_idx, column=1).alignment = CENTER_ALIGN
        
        ws.cell(row=row_idx, column=2, value=col.get('column_name', '')).border = BORDER
        ws.cell(row=row_idx, column=3, value=col.get('data_type', '')).border = BORDER
        ws.cell(row=row_idx, column=3).alignment = CENTER_ALIGN
        
        # 길이
        length = col.get('max_length') or col.get('character_maximum_length') or ''
        if length == -1:
            length = 'MAX'
        ws.cell(row=row_idx, column=4, value=length).border = BORDER
        ws.cell(row=row_idx, column=4).alignment = CENTER_ALIGN
        
        # PK
        pk = '✓' if col.get('is_primary_key') else ''
        ws.cell(row=row_idx, column=5, value=pk).border = BORDER
        ws.cell(row=row_idx, column=5).alignment = CENTER_ALIGN
        
        # NULL
        nullable = 'Y' if col.get('is_nullable') else 'N'
        ws.cell(row=row_idx, column=6, value=nullable).border = BORDER
        ws.cell(row=row_idx, column=6).alignment = CENTER_ALIGN
        
        # 기본값
        default = col.get('default_value') or col.get('column_default') or ''
        ws.cell(row=row_idx, column=7, value=str(default)[:50]).border = BORDER
        
        # 설명
        desc = col.get('description') or col.get('column_description') or ''
        ws.cell(row=row_idx, column=8, value=desc).border = BORDER
        
        # 비고
        ws.cell(row=row_idx, column=9, value='').border = BORDER
    
    set_column_widths(ws, {
        'A': 6, 'B': 25, 'C': 15, 'D': 8, 
        'E': 6, 'F': 6, 'G': 20, 'H': 30, 'I': 20
    })


@router.get("/tables/{server_id}/{db_name}")
async def get_tables_for_export(
    server_id: int,
    db_name: str,
    db: Session = Depends(get_db),
    user: User = Depends(require_login)
):
    """테이블 목록 조회 (정의서 내보내기용)"""
    server_service = ServerService(db)
    server = server_service.get_server(server_id)
    
    if not server:
        return JSONResponse({"error": "서버를 찾을 수 없습니다"}, status_code=404)
    
    driver = get_driver(server)
    tables = driver.get_tables(db_name)
    
    # 컬럼 수 추가
    for table in tables:
        try:
            columns = driver.get_table_columns(db_name, table['table_name'])
            table['column_count'] = len(columns)
            print(f"테이블 {table['table_name']}: 컬럼 {len(columns)}개")  # 추가
        except Exception as e:
            print(f"컬럼 조회 실패 {table['table_name']}: {e}")  # 추가
            table['column_count'] = 0
    
    return {"tables": tables}


@router.get("/download/server/{server_id}")
async def download_server_schema(
    server_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_login)
):
    """서버 전체 DB 정의서 다운로드"""
    server_service = ServerService(db)
    server = server_service.get_server(server_id)
    
    if not server:
        return JSONResponse({"error": "서버를 찾을 수 없습니다"}, status_code=404)
    
    driver = get_driver(server)
    databases = driver.get_databases()
    
    print(f"\n{'='*60}")
    print(f"📊 테이블 정의서 생성 시작: {server.server_name}")
    print(f"   총 {len(databases)}개 DB 처리 예정")
    print(f"{'='*60}")
    
    wb = Workbook()
    ws_toc = wb.active
    ws_toc.title = "목차"
    
    # 목차 시트 - 제목
    ws_toc.merge_cells('A1:H1')
    ws_toc['A1'] = "테이블 정의서"
    ws_toc['A1'].font = Font(bold=True, size=20)
    ws_toc['A1'].alignment = CENTER_ALIGN
    ws_toc.row_dimensions[1].height = 40
    
    # 서버 정보
    ws_toc['A3'] = "서버"
    ws_toc['B3'] = f"{server.server_name} ({server.host}:{server.port})"
    ws_toc['A3'].font = Font(bold=True)
    
    ws_toc['A4'] = "생성일"
    ws_toc['B4'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws_toc['A4'].font = Font(bold=True)
    
    ws_toc['A5'] = "생성자"
    ws_toc['B5'] = user.username
    ws_toc['A5'].font = Font(bold=True)
    
    # 목차 헤더
    ws_toc['A7'] = "■ 전체 테이블 목록"
    ws_toc['A7'].font = SUBTITLE_FONT
    
    toc_headers = ['No', 'DB명', '테이블명', '테이블설명', '컬럼 수', '행 수', '용량(MB)', '시트명']
    for col_idx, header in enumerate(toc_headers, 1):
        cell = ws_toc.cell(row=8, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = BORDER
    
    toc_row = 9
    table_no = 1
    total_db_count = len(databases)
    processed_db_count = 0
    
    # 각 DB 처리
    for db_idx, db_info in enumerate(databases, 1):
        db_name = db_info['db_name']
        print(f"\n[{db_idx}/{total_db_count}] DB 처리 중: {db_name}")
        
        try:
            tables = driver.get_tables(db_name)
            if not tables:
                print(f"  └─ 테이블 없음, 건너뜀")
                continue
            
            print(f"  └─ 테이블 {len(tables)}개 발견")
            processed_db_count += 1
            
            # DB 시트 생성 (31자 제한)
            sheet_name = db_name[:31]
            base_name = sheet_name
            counter = 1
            while sheet_name in wb.sheetnames:
                sheet_name = f"{base_name[:28]}_{counter}"
                counter += 1
            
            ws_db = wb.create_sheet(title=sheet_name)
            
            # DB 시트 제목
            ws_db['A1'] = f"DB: {db_name}"
            ws_db['A1'].font = Font(bold=True, size=16)
            ws_db['A2'] = f"테이블 수: {len(tables)}"
            ws_db['A2'].font = Font(size=10, color="666666")
            
            # DB 시트 헤더
            db_headers = ['No', '테이블명', '테이블설명', '컬럼명', '데이터타입', '길이', 'PK', 'NULL', '기본값', '설명', '비고']
            for col_idx, header in enumerate(db_headers, 1):
                cell = ws_db.cell(row=4, column=col_idx, value=header)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = CENTER_ALIGN
                cell.border = BORDER
            
            db_row = 5  # 데이터 시작 행
            col_no = 1
            
            for tbl_idx, table in enumerate(tables, 1):
                table_name = table['table_name']
                table_desc = table.get('description') or table.get('table_description') or ''
                
                # 컬럼 정보 조회
                try:
                    columns = driver.get_table_columns(db_name, table_name)
                except:
                    columns = []
                
                # 목차에 추가
                ws_toc.cell(row=toc_row, column=1, value=table_no).border = BORDER
                ws_toc.cell(row=toc_row, column=1).alignment = CENTER_ALIGN
                ws_toc.cell(row=toc_row, column=2, value=db_name).border = BORDER
                ws_toc.cell(row=toc_row, column=3, value=table_name).border = BORDER
                ws_toc.cell(row=toc_row, column=4, value=table_desc).border = BORDER
                ws_toc.cell(row=toc_row, column=5, value=len(columns)).border = BORDER
                ws_toc.cell(row=toc_row, column=5).alignment = CENTER_ALIGN
                ws_toc.cell(row=toc_row, column=6, value=table.get('row_count', 0)).border = BORDER
                ws_toc.cell(row=toc_row, column=6).alignment = CENTER_ALIGN
                ws_toc.cell(row=toc_row, column=7, value=round(table.get('size_mb', 0), 2)).border = BORDER
                ws_toc.cell(row=toc_row, column=7).alignment = CENTER_ALIGN
                
                # 시트 링크
                link_cell = ws_toc.cell(row=toc_row, column=8, value=sheet_name)
                link_cell.hyperlink = f"#'{sheet_name}'!A{db_row}"
                link_cell.font = Font(color="0563C1", underline="single")
                link_cell.border = BORDER
                
                toc_row += 1
                table_no += 1
                
                # DB 시트에 컬럼 데이터 추가
                for col in columns:
                    ws_db.cell(row=db_row, column=1, value=col_no).border = BORDER
                    ws_db.cell(row=db_row, column=1).alignment = CENTER_ALIGN
                    
                    ws_db.cell(row=db_row, column=2, value=table_name).border = BORDER
                    ws_db.cell(row=db_row, column=3, value=table_desc).border = BORDER
                    ws_db.cell(row=db_row, column=4, value=col.get('column_name', '')).border = BORDER
                    ws_db.cell(row=db_row, column=5, value=col.get('data_type', '')).border = BORDER
                    ws_db.cell(row=db_row, column=5).alignment = CENTER_ALIGN
                    
                    length = col.get('max_length') or ''
                    if length == -1:
                        length = 'MAX'
                    ws_db.cell(row=db_row, column=6, value=length).border = BORDER
                    ws_db.cell(row=db_row, column=6).alignment = CENTER_ALIGN
                    
                    pk = '✓' if col.get('is_primary_key') else ''
                    ws_db.cell(row=db_row, column=7, value=pk).border = BORDER
                    ws_db.cell(row=db_row, column=7).alignment = CENTER_ALIGN
                    
                    nullable = 'Y' if col.get('is_nullable') else 'N'
                    ws_db.cell(row=db_row, column=8, value=nullable).border = BORDER
                    ws_db.cell(row=db_row, column=8).alignment = CENTER_ALIGN
                    
                    ws_db.cell(row=db_row, column=9, value=str(col.get('default_value') or col.get('column_default') or '')[:50]).border = BORDER
                    ws_db.cell(row=db_row, column=10, value=col.get('description') or '').border = BORDER
                    ws_db.cell(row=db_row, column=11, value='').border = BORDER  # 비고
                    
                    db_row += 1
                    col_no += 1
                
                # 10개마다 또는 마지막 테이블일 때 진행상황 출력
                if tbl_idx % 10 == 0 or tbl_idx == len(tables):
                    print(f"  └─ 테이블 처리 중: {tbl_idx}/{len(tables)}")
            
            # DB 시트 컬럼 너비 설정
            set_column_widths(ws_db, {
                'A': 8, 'B': 25, 'C': 25, 'D': 25, 'E': 15, 'F': 8,
                'G': 6, 'H': 6, 'I': 20, 'J': 30, 'K': 20
            })
            
        except Exception as e:
            print(f"  └─ ❌ 처리 실패: {e}")
    
    # 목차 시트 컬럼 너비 설정
    set_column_widths(ws_toc, {
        'A': 8, 'B': 20, 'C': 25, 'D': 25, 'E': 10, 'F': 12, 'G': 12, 'H': 20
    })
    
    print(f"\n{'='*60}")
    print(f"✅ 엑셀 생성 완료!")
    print(f"   처리된 DB: {processed_db_count}개")
    print(f"   총 테이블: {table_no - 1}개")
    print(f"{'='*60}\n")
    
    # 파일 저장
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"{server.server_name}_테이블정의서_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    # 활동 로그 기록
    log_download_schema_all(
        db=db,
        user_id=user.id,
        username=user.username,
        server_id=server_id,
        server_name=server.server_name,
        db_count=len(databases),
        filename=filename
    )
    
    response = StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"}
    )
    response.set_cookie(key="download_complete", value="true", max_age=5)
    return response

@router.get("/download/db/{server_id}/{db_name}")
async def download_db_schema(
    server_id: int,
    db_name: str,
    tables: Optional[str] = Query(None, description="쉼표로 구분된 테이블 목록"),
    db: Session = Depends(get_db),
    user: User = Depends(require_login)
):
    """단일 DB 정의서 다운로드"""
    server_service = ServerService(db)
    server = server_service.get_server(server_id)
    
    if not server:
        return JSONResponse({"error": "서버를 찾을 수 없습니다"}, status_code=404)
    
    driver = get_driver(server)
    all_tables = driver.get_tables(db_name)
    
    # 테이블 필터링
    if tables:
        selected_tables = tables.split(',')
        all_tables = [t for t in all_tables if t['table_name'] in selected_tables]
    
    # 컬럼 수 추가
    for table in all_tables:
        try:
            columns = driver.get_table_columns(db_name, table['table_name'])
            table['column_count'] = len(columns)
        except:
            table['column_count'] = 0
    
    wb = Workbook()
    
    # 목차 생성
    create_db_toc_sheet(wb, f"{server.server_name} ({server.host}:{server.port})", db_name, all_tables, user.username)
    
    # 테이블 시트 생성
    for table in all_tables:
        try:
            columns = driver.get_table_columns(db_name, table['table_name'])
            create_table_sheet(wb, db_name, table['table_name'], columns, table)
        except Exception as e:
            print(f"테이블 {table['table_name']} 처리 실패: {e}")
    
    # 파일 저장
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"{db_name}_테이블정의서_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    # 활동 로그 기록
    log_download_schema(
        db=db,
        user_id=user.id,
        username=user.username,
        server_id=server_id,
        db_name=db_name,
        table_count=len(all_tables),
        filename=filename
    )
    
    response = StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"}
    )
    response.set_cookie(key="download_complete", value="true", max_age=5)
    return response